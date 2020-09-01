from openpyxl import load_workbook
fname = "12312312312.xlsx" #양식
#fname = "C:\\Users\\owner\\Documents\\example.xlsx"
wb = load_workbook(fname)
#시트 이름 리스트
sheets = wb.sheetnames
smax = len(sheets)

cform=[]

#시트별 루프
for i in range(smax):
  sname = sheets[i] #시트명
  ws = wb[sname]
  '''HERE 사이즈 필요'''
  #개선점: 가장 긴 첫 줄 기준으로 길이 변경
  fline = ws['A1:H1'] #tuple
  tmp = []
  '''HERE 사이즈 필요'''
  for i in range(8): #적당한 값 대입 가능
    tmp.append(str(fline[0][i].value))
  cform.append(tmp)
print(cform)