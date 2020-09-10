#요구사항 바뀜
#import getCForm
from openpyxl import load_workbook
#파일 리스트 가져오기

from os import listdir, getcwd
path_dir = getcwd()
file_list = listdir(path_dir)
fnames = []
for i in file_list:
    if i[-4:] == 'xlsx' and i[:2] != '~$' and (len(i.split('_')[0]) > 5):
        fnames.append(i)

def isNone(str=None):
    if str is None: return True
    for c in str:
        if c != ' ': return False
    return True

def whatRC(row, col):
    return str(chr(ord('A')+col-1))+str(row)

#첫 줄 비교
def isEqui(t1, t2):
    for i in range(len(t2)):
        if t1[i] != str(t2[i].value):
            #연도용
            if '-12-31' in t1[i]:
                if t1[i][:4] != str(t2[i].value)[:4]:
                    return False
            elif t1[i] != 'wildcard':
                return False
    return True

#cform = getCForm.getCForm() #extract
#hard coding
cform = [[
    '기업체명', '사업자번호', '대표자명', '설립형태', '기업형태', '전화번호', '홈페이지', '결산월', '주소',
    'None', '업종(10차)', '업종(9차)', '주요제품(상품)', '무역업허가번호', '주채권기관', '휴폐업정보',
    'None', '영문기업명', '법인(주민)번호', '종업원수', '설립일자', '기업규모', '팩스번호', '이메일',
    '기업공개일자', 'None', 'None', 'None', 'None', 'None', '소속그룹', '당좌거래은행',
    '법인등기정보', 'None'
], ['결산기준일자', '총자산', '납입자본금', '자본총계', '매출액', '영업이익', '순이익'], ['연혁일자 ▲▼', '내용'],
         ['내용'],
         [
             '구분', '상시종업원수', 'None', 'None', 'None', 'None', '평균근속년수', '연급여총액',
             '1인평균'
         ],
         [
             '구분', '상시종업원수', 'None', 'None', 'None', 'None', '평균근속년수', '연급여총액',
             '1인평균'
         ],
         [
             '구분', '상시종업원수', 'None', 'None', 'None', 'None', '평균근속년수', '연급여총액',
             '1인평균'
         ],
         [
             '구분', '상시종업원수', 'None', 'None', 'None', 'None', '평균근속년수', '연급여총액',
             '1인평균'
         ],
         [
             '구분', '상시종업원수', 'None', 'None', 'None', 'None', '평균근속년수', '연급여총액',
             '1인평균'
         ],
         [
             '구분', '상시종업원수', 'None', 'None', 'None', 'None', '평균근속년수', '연급여총액',
             '1인평균'
         ], ['번호', '구분', '사업자번호', '사업장명', '주생산품', '주소', '전화번호'],
         [
             '설비명', 'wildcard', 'None', 'None', 'wildcard', 'None', 'None',
             '2015-12-31 00:00:00', 'None', 'None', 'None', 'None', 'None',
             'None'
         ],
         [
             '설비명', 'wildcard', 'None', 'None', 'wildcard', 'None', 'None',
             '2016-12-31 00:00:00', 'None', 'None'
         ],
         [
             '설비명', 'wildcard', 'None', 'None', 'wildcard', 'None', 'None',
             '2017-12-31 00:00:00', 'None', 'None'
         ],
         [
             '설비명', 'wildcard', 'None', 'None', 'wildcard', 'None', 'None',
             '2018-12-31 00:00:00', 'None', 'None'
         ],
         [
             '설비명', 'wildcard', 'None', 'None', 'wildcard', 'None', 'None',
             '2019-12-31 00:00:00', 'None', 'None'
         ],
         [
             '설비명', 'wildcard', 'None', 'None', 'wildcard', 'None', 'None',
             '2020-12-31 00:00:00', 'None', 'None'
         ], ['사업부문', '주요제품(품목)', '매출액구성비(%)'],
         ['사업부문', '주요제품(품목)', '매출액구성비(%)'], ['사업부문', '주요제품(품목)', '매출액구성비(%)'],
         ['사업부문', '주요제품(품목)', '매출액구성비(%)'],
         [
             '사업부문', '주요제품(품목)', '매출액구성비(%)', 'None', 'None', 'None', 'None',
             'None', 'None', 'None', 'None', 'None'
         ], ['사업부문', '주요제품(품목)', '매출액구성비(%)'],
         ['기업명', '사업자번호', '대표자명', '거래비중', '결산년도', '자본금', '자산총계', '매출액', '순이익'],
         ['기업명', '사업자번호', '대표자명', '거래비중', '결산년도', '자본금', '자산총계', '매출액', '순이익'],
         ['기업명', '사업자번호', '대표자명', '거래비중', '결산년도', '자본금', '자산총계', '매출액', '순이익'],
         [
             '기업명', '사업자번호', '대표자명', '거래비중', '결산년도', '자본금', '자산총계', '매출액',
             '순이익', 'None', 'None', 'None', 'None'
         ],
         [
             '기업명', '사업자번호', '대표자명', '거래비중', '결산년도', '자본금', '자산총계', '매출액',
             '순이익', 'None', 'None', 'None', 'None', 'None', 'None', 'None',
             'None', 'None', 'None', 'None', 'None'
         ],
         ['기업명', '사업자번호', '대표자명', '거래비중', '결산년도', '자본금', '자산총계', '매출액', '순이익'],
         ['기업명', '사업자번호', '대표자명', '거래비중', '결산년도', '자본금', '자산총계', '매출액', '순이익'],
         ['기업명', '사업자번호', '대표자명', '거래비중', '결산년도', '자본금', '자산총계', '매출액', '순이익'],
         ['기업명', '사업자번호', '대표자명', '거래비중', '결산년도', '자본금', '자산총계', '매출액', '순이익'],
         ['기업명', '사업자번호', '대표자명', '거래비중', '결산년도', '자본금', '자산총계', '매출액', '순이익'],
         ['기업명', '사업자번호', '대표자명', '거래비중', '결산년도', '자본금', '자산총계', '매출액', '순이익'],
         ['기업명', '사업자번호', '대표자명', '거래비중', '결산년도', '자본금', '자산총계', '매출액', '순이익'],
         ['등급', '평가(산출)일자', '재무기준일자', '등급구분'],
         ['계정명', 'wildcard', 'wildcard', '2015-12-31 00:00:00'],
         ['계정명', 'wildcard', 'wildcard', '2016-12-31 00:00:00'],
         ['계정명', 'wildcard', 'wildcard', '2017-12-31 00:00:00'],
         ['계정명', 'wildcard', 'wildcard', '2018-12-31 00:00:00'],
         ['계정명', 'wildcard', 'wildcard', '2019-12-31 00:00:00'],
         ['계정명', 'wildcard', 'wildcard', '2020-12-31 00:00:00'],
         ['계정명', 'wildcard', 'wildcard', '2015-12-31 00:00:00'],
         ['계정명', 'wildcard', 'wildcard', '2016-12-31 00:00:00'],
         ['계정명', 'wildcard', 'wildcard', '2017-12-31 00:00:00'],
         ['계정명', 'wildcard', 'wildcard', '2018-12-31 00:00:00'],
         ['계정명', 'wildcard', 'wildcard', '2019-12-31 00:00:00'],
         ['계정명', 'wildcard', 'wildcard', '2020-12-31 00:00:00'],
         ['계정명', 'wildcard', 'wildcard', '2015-12-31 00:00:00'],
         ['계정명', 'wildcard', 'wildcard', '2016-12-31 00:00:00'],
         ['계정명', 'wildcard', 'wildcard', '2017-12-31 00:00:00'],
         ['계정명', 'wildcard', 'wildcard', '2018-12-31 00:00:00'],
         ['계정명', 'wildcard', 'wildcard', '2019-12-31 00:00:00'],
         ['계정명', 'wildcard', 'wildcard', '2020-12-31 00:00:00']]

SHEETNO = 54
infoStr = ''
sinfoStr = ''  #상세
rn = '\n'

#개선점: 몇천 개 한 번에 (스트링 버퍼 주의 끊어서 돌리기)
#개선점: 나중에 읽어오도록 고치기,
#개선점: tkinter로 표시
#개선점: 어디가 의심스러운지 표시

fo = open('report.txt', "w+")
fo1 = open('상세report.txt', "w+")

#output
susFileNames=[]
susFileSus=[]

for fname in fnames:
    infoStr = infoStr + '파일 ' + fname + ': ' + rn
    wb = load_workbook(fname)

    emptySheetNames = []  #빈 시트 이름 리스트
    milSheetNames = []  #밀린 시트 이름 리스트
    susSheetNames = []  #첫 줄/형식 의심스러운 시트 이름 리스트
    errStr = ''

    #시트 이름 리스트
    sheets = wb.sheetnames
    smax = len(sheets)
    if smax != SHEETNO: errStr = errStr + "시트 수: " + smax + rn

    #시트별 루프
    for n in range(smax):
        sname = sheets[n]  #시트명
        ws = wb[sname]
        fline = ws[1]  #tuple
        sline = ws[2]  #second line(empty check)
        if n == 0:  #시트1
            fline = ws['a']
            sline = ws['d']
           
        #빈 시트 체크
        if isNone(fline[0].value):
            #밀린 시트 처리 방법 찾아야함
            if ws[whatRC(ws.max_row, ws.max_column)].value is not None: #' '여도 밀렸음
                milSheetNames.append(sname)
                susSheetNames.append(sname)
            else:
                #빈 시트
                emptySheetNames.append(sname)
            continue  #해당 시트에서 이하 코드 필요 없음
        else:  #문제없음
            pass

        #첫 페이지와 기업체명 체크
        if n == 0:  #첫 시트
            fline += sline
            if not isEqui(cform[0], fline):
                susSheetNames.append(sname)
            fcname = str(ws['C1'].value)  #'(주)' 빼야함
            cname = fname.split('_')[0][5:]
            if cname not in fcname:
                errStr = errStr + '파일명이 기업체명과 다릅니다: 파일명-' + fcname + ', 기업체명-' + cname + rn
                susSheetNames.append('1.의 파일명')
            if str(ws['C5'].value) == '개인사업자':
              errStr += '개인사업자입니다.'+rn
              susSheetNames.append('1.의 기업형태: 개인사업자')
              continue #작동하나?
            if '폐업' in str(ws['c16'].value):
              errStr += '폐업자입니다.'+rn
              susSheetNames.append('1.의 휴폐업정보: 폐업')
              continue

        #첫 줄 체크
        #일반적인 범위
        else:
            if not isEqui(cform[n], fline):
                #err msg
                susSheetNames.append(sname)
            #연도 체크 넣는다면 여기
        #넘어가는 범위
    wb.close()

    if len(susSheetNames) > 0:
        susFileNames.append(fname)
        susFileSus.append(susSheetNames)
        errStr = errStr + '첫 줄이 의심스러운 시트 리스트: ' + str(susSheetNames) + rn
    if len(milSheetNames) > 0:
        errStr = errStr + '붙여넣기 위치가 잘못된 시트 리스트: ' + str(milSheetNames) + rn
    if len(emptySheetNames) > 0:
        errStr = errStr + '빈 시트 리스트: ' + str(emptySheetNames) + rn

    if errStr == '':
        infoStr += 'OK' + rn
    else:
        infoStr = infoStr + errStr + rn

fo1.write(infoStr)
fo1.close()

fileInfoStr = ''
if len(susFileNames) == 0:
    fileInfoStr = '모든 파일 OK'
else:
    for i in range(len(susFileNames)):
        fileInfoStr += '파일 ' + susFileNames[i] + rn
        for j in susFileSus[i]:
            fileInfoStr += '의심 시트: ' + j + rn
        fileInfoStr += rn

fo.write(fileInfoStr)
fo.close()
