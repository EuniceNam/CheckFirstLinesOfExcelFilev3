from openpyxl import load_workbook

def getCForm():

  fname = "0123.포맷_테스트.xlsx" #양식
  wb = load_workbook(fname)
  #시트 이름 리스트
  sheets = wb.sheetnames
  smax = len(sheets)

  cform=[]

  #시트별 루프
  for i in range(smax):
    sname = sheets[i] #시트명
    ws = wb[sname]
    fline = ws[1] #tuple
    if i == 0:
      fline = ws['a']+ws['d'] #첫 시트
    tmp = []
    for i in range(len(fline)):
      tmp.append(str(fline[i].value))
    cform.append(tmp)
  fo = open('cform.txt', 'w+')
  fo.write(str(cform))
  fo.close()
  return cform
