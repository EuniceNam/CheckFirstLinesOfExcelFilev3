from openpyxl import load_workbook
from os import listdir, getcwd
import json
import Part #isEqui(), lst(), cform


#config: 시군, 검사할 범위(를 입력받을 건지와 그 범위), 검사 방법(순서대로, 있는 것만)
config={}
with open('config.json', encoding='UTF8') as j:
    config=json.load(j)
INDEX, SH_NAME, CORPNS, CORPNF, partS, city = config['listFileName'], config['listSheetName'], config['startNo'], config['finNo'], config['fileStartNo'], config['city']
if config['partOnly']=='True': lst = Part.lst(INDEX, sheetname=SH_NAME) #54시트 파일이 있는 경우만 숫자->리스트와 비교 (수정본만 검수할 때). 리스트는 전체를 가져옴.
elif config['rangeInput']=='True': #범위 나눠서 테스트할 때
    CORPNS, CORPNF = input("비교할 목록 시작번호 끝번호: ").split(' ',1)
    partS = int(input("폴더 내 파일 시작번호: "))
    lst = Part.lst(INDEX, CORPNS=CORPNS, CORPNF=CORPNF, sheetname=SH_NAME)
else: #기본
    lst = Part.lst(INDEX, CORPNS=CORPNS, CORPNF=CORPNF, sheetname=SH_NAME)

#폴더 내 파일 목록 가져오기
path_dir = getcwd()+'/'+config['path']
file_list = listdir(path_dir)
fnames = []
for i in file_list:
    if i[-4:] == 'xlsx' and i[:2] != '~$' and (len(i.split('_')[0]) > 5) and i[0] in '0123456789':
        fnames.append(i)

if partS != 'default' and partS != '':
    if CORPNS == 'default' or CORPNS == '': CORPNS=1
    else: CORPNS = int(CORPNS)
    if CORPNF == 'default' or CORPNF == '': CORPNF = len(lst)
    else: CORPNF = int(CORPNF)
    fnames = fnames[CORPNS-int(partS):CORPNF-int(partS)+1] 

def isNone(string): #빈(띄어쓰기) 스트링
    if string is None: return True #useless?
    for c in str(string):
        if c != ' ': return False
    return True

def whatRC(row, col): #좌표 숫자 -> 'A1' 등 엑셀 좌표로 변환
    return str(chr(ord('A')+col-1))+str(row)

SHEETNO = 54
infoStr = ''
sinfoStr = ''  #상세
rn = '\n'
tab='\t'

fo = open(config['city']+'_report.txt', "w+")
fo1 = open(config['city']+'_상세report.txt', "w+")
susFileNames = [] #의심스러운 파일 이름
susFileSus = [] #의심스러운 파일에 해당하는 의심
filecnt=0

print("Now examining...")
errtmp=-2
ntmp=0
try:
    for fname in fnames:
        filecnt+=1
        #진도 프롬프트에 띄우기
        if filecnt%100 == 0:
            print(str(filecnt)+'th file '+fname)
        errStr=''
        infoStr = fname + ': ' + rn
        errtmp=-1
        #print('fname: "'+fname+'"')
        try:
            wb = load_workbook(path_dir+'/'+fname)
        except:
            fo.write("파일 "+fname+"을 열 수 없음 -> 직접 검수 필요\n") #보통 datetime 에러, 몇천 개에 하나 정도
            fo1.write(fname+": 열 수 없음\n")
            continue #이 경우 이하 코드 못 씀
        errtmp=-0.5
        emptySheetNames = []  #빈 시트 이름 리스트
        milSheetNames = []  #밀린 시트 이름 리스트
        susSheetNames = []  #첫 줄/형식 의심스러운 시트 이름 리스트

        #시트 이름 리스트
        sheets = wb.sheetnames
        smax = len(sheets)
        if smax != SHEETNO: 
            errStr = errStr + "시트 수: " + str(smax) + ", 검사 못하고 지나감" + rn
            fo1.write(errStr)
            continue #이하 코드 필요 없음
        #문제확인용
        errtmp, ntmp=0,0
        #시트별 루프
        for n in range(smax):
            errtmp, ntmp=1,n #문제확인용
            lnum, lcname, lcid = '', '', '' #용인시 리스트의 숫자, 업체명, 사업자번호
            fnum, ncname, fcname, fcid = '','','','' #파일명의 숫자, 파일명의 업체명, 파일내의 업체명, 파일내의 사업자번호
            sname = sheets[n]  #시트명
            #print('sname: '+sname)
            ws = wb[sname]
            fline = ws[1]  #first line
            sline = ws[2]  #second line
            errtmp=2
            if n == 0:  #시트1
                fline = ws['a']
                sline = ws['d']
                [lnum, lcname, lcid] = lst[filecnt-1] #용인시 리스트의 숫자, 업체명, 사업자번호
            errtmp=2.5              
            #빈 시트 체크
            if isNone(fline[0].value):
                #밀린 시트
                errtmp = 2.6
                if ws[whatRC(ws.max_row, ws.max_column)].value is not None: #' '여도 밀렸음
                    milSheetNames.append(sname)
                    susSheetNames.append(sname+' 붙여넣기 위치 밀림')
                else:
                    #빈 시트
                    errtmp = 2.7
                    emptySheetNames.append(sname)
                    errtmp = 2.8
                continue  #해당 시트에서 이하 코드 필요 없음
            else:  #문제없음
                errtmp=2.9
                pass
            errtmp=3
            #첫 시트
            if n == 0: 
                #형식 체크
                if not Part.isEqui(Part.cform[0][0], fline) or not Part.isEqui(Part.cform[0][1], sline):
                    susSheetNames.append(sname+Part.printRow(fline))
                fcname, fcid = str(ws['C1'].value), str(ws['C2'].value)  #앞뒤'(주)' 빼야함
                if fcname[-1] == ' ': fcname=fcname[:-1]
                if fcname[:3] == '(주)': fcname=fcname[3:]
                if fcname[-3:] == '(주)': fcname=fcname[:-3]
                if fcname[0] == '㈜': fcname=fcname[1:]
                if fcname[-1] == '㈜': fcname=fcname[:-1]
                fcname = fcname.replace('  ', ' ') #가끔 띄어쓰기 2개로 읽히는 칸 있음
                errtmp=4
                ncname = fname.split('_')[0]
                errtmp=4.1
                if '.' in ncname: fnum, ncname = ncname.split('.', 1)
                else: #파일명 오류 중 하나, 이름 검사에서 걸리도록 일부러 겹치게 해둠
                    fnum, ncname = ncname[:4], ncname[4:]
                errtmp=4.2
                if '(주)' in ncname or '㈜' in ncname: susSheetNames.append('(주).파일명')
                #비교할 것: lnum, fnum, lcname, ncname, fcname, lcid, fcid
                #일부검수용
                if config['partOnly']=='True': [lnum, lcname, lcid] = lst[int(fnum)-1] #일부 숫자만 검수할 때
                errtmp=4.3
                if lnum != fnum: susSheetNames.append('1.의 파일명 숫자')
                tmpname=''
                if lcname != ncname or lcname != fcname or ncname != fcname: tmpname+='("리스트", "파일명", "1.기업명"") = ("'+lcname+'", "'+ncname+'", "'+fcname+'"")'
                if tmpname!='': susSheetNames.append('1.등의 업체명: '+tmpname)
                if lcid != fcid: susSheetNames.append('1.의 사업자번호: '+fcid)
                errtmp=4.4
                if '개인사업자' in str(ws['C5'].value):
                  susSheetNames.append('1.의 기업형태: 개인사업자')
                if '폐업' in str(ws['c16'].value):
                  susSheetNames.append('1.의 휴폐업정보: 폐업')
                if '휴업' in str(ws['c16'].value):
                  susSheetNames.append('1.의 휴폐업정보: 휴업')
                if config['city'] not in str(ws['c9'].value):
                  susSheetNames.append('1.의 주소: '+str(ws['c9'].value))

            #2~54시트: 첫 줄 체크
            else:
                errtmp=5
                if not Part.isEqui(Part.cform[n], fline):
                    #err msg
                    susSheetNames.append(sname)
        
        errtmp=6
        wb.close()
        errtmp=7
        #output
        if len(emptySheetNames) >= SHEETNO: 
            susSheetNames.append('모든 시트가 빈 파일')
        if len(susSheetNames) > 0:
            susFileNames.append(fname)
            susFileSus.append(susSheetNames)
            errStr+= '-첫 줄이 의심스러운 시트: '+rn
            for i in susSheetNames:
                errStr+=tab+i+rn
        if len(milSheetNames) > 0:
            errStr+= '-위치 밀린 시트:'+rn
            for i in milSheetNames:
                errStr+=tab+i+rn
        if len(emptySheetNames) > 0:
            errStr+='-빈 시트: '+rn
            for i in emptySheetNames:
                errStr+=tab+i+rn
        if errStr == '':
            infoStr += 'OK' + rn
        else:
            infoStr += errStr + rn

        fo1.write(infoStr)
        errtmp=10
except:
    fo.write("err"+rn)
    fo1.write("err"+rn)
    print("error")
    print("errtmp: "+str(errtmp)+', ntmp: '+str(ntmp))
finally:
    fo.write("filecnt: "+str(filecnt)+rn)
    fo1.write("filecnt: "+str(filecnt)+rn)
    fo1.close()

#fo1.write(infoStr)
#fo1.close()

fileInfoStr = ''
if len(susFileNames) == 0:
    fileInfoStr = '모든 파일 OK'
else:
    for i in range(len(susFileNames)):
        fileInfoStr += '파일 ' + susFileNames[i] + rn
        for j in susFileSus[i]:
            fileInfoStr += '의심 시트: ' + j + rn
        fileInfoStr += rn

fo.write(fileInfoStr)
fo.close()

print("finished")
