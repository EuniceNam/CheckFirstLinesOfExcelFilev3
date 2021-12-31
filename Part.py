from openpyxl import load_workbook
#시군별 다운로드 목록에서 번호, 업체명, 사업자번호 추출 및 처리 #어차피 많아봤자 만 개 조금 넘으니 리스트 하나에 넣음
def lst(fname, CORPNS=1, CORPNF=None, sheetname=None):
    if sheetname == None or sheetname=='' or sheetname=='default': sheet = load_workbook(fname).active
    else: sheet = load_workbook(fname)[sheetname]
    tnums, tcnames, tcids = sheet['A'], sheet['B'], sheet['C']
    if CORPNS=="default" or CORPNS==None: CORNPS=1
    if CORPNF=="default" or CORPNF==None: CORPNF=len(tnums)-1
    lst=[]
    for i in range(int(CORPNS), int(CORPNF)+1):
        num=str(tnums[i].value)
        tmp=''
        for j in range(4-len(num)):
            tmp+='0'
        num=tmp+num
        cname = str(tcnames[i].value)
        if cname[:3] == '(주)': cname=cname[3:]
        if cname[-3:] == '(주)': cname=cname[:-3]
        if cname[0] == '㈜': cname=cname[1:]
        if cname[-1] == '㈜': cname=cname[:-1]
        cname = cname.replace('  ', ' ') #농업회사법인에서 띄어쓰기가 2개로 읽히는 경우가 꽤 있음
        if cname[-1] == ' ': cname = cname[:-1]
        cid = str(tcids[i].value)
        if cid[-1] == ' ': cid=cid[:-1]
        if len(cid)>5:
            if cid[3]!= '-': cid=cid[:3]+'-'+cid[3:] 
            if cid[6]!= '-': cid=cid[:6]+'-'+cid[6:]
        lst.append([num, cname, cid])
    return lst

def printRow(cells): #문제 확인용
    ret=''
    for i in cells:
        ret+=','+str(i.value)
    return ret[1:]

def isEqui(t1, t2): #지정된 형식 t1과 엑셀 영역 t2 비교
    #잘못 복붙했거나(연도 포함) 영역 외 셀에 띄어쓰기가 입력된 경우 등 때문
    tmp = len(t1)-len(t2)
    mn = len(t1) #비교할 영역 크기
    #남는 영역이 있다면 모두 None이어야함
    if tmp>0: 
        mn=len(t2) #짧은쪽
        for i in t1[-tmp:]:
            if i != 'None': 
                return False
    elif tmp<0: 
        for  i in t2[tmp:]:
            if str(i.value) != 'None': 
                return False
    #여기 빈 시트는 입력으로 안 들어와야함
    for i in range(mn):
        if t1[i] != str(t2[i].value):
            #연도용 #월일만 다른 경우
            if len(t1[i]) >= 10 and t1[i][4:10] == '-12-31': #연도일 때
                if t1[i][:4] != str(t2[i].value)[:4]:
                    return False
            elif t1[i] != 'wildcard':
                return False
    return True

#기타사항: 2.경영규모에서 cform[1][4]가 매출액 대신 영업수익인 경우 fo에 문제로 기록되나 문제없음
cform = [
         [['기업체명', '사업자번호', '대표자명', '설립형태', '기업형태', '전화번호', '홈페이지', '결산월', '주소', 'None', '업종(10차)', '업종(9차)', '주요제품(상품)', '무역업허가번호', '주채권기관', '휴폐업정보', 'None'], 
         ['영문기업명', '법인(주민)번호', '종업원수', '설립일자', '기업규모', '팩스번호', '이메일', '기업공개일자', 'None', 'None', 'None', 'None', 'None', '소속그룹', '당좌거래은행', '법인등기정보', 'None']], 
         ['결산기준일자', '총자산', '납입자본금', '자본총계', '매출액', '영업이익', '순이익'], 
         ['연혁일자 ▲▼', '내용'],
         ['내용'],
         ['구분', '상시종업원수', 'None', 'None', 'None', 'None', '평균근속년수', '연급여총액', '1인평균'],
         ['구분', '상시종업원수', 'None', 'None', 'None', 'None', '평균근속년수', '연급여총액', '1인평균'],
         ['구분', '상시종업원수', 'None', 'None', 'None', 'None', '평균근속년수', '연급여총액', '1인평균'],
         ['구분', '상시종업원수', 'None', 'None', 'None', 'None', '평균근속년수', '연급여총액', '1인평균'],
         ['구분', '상시종업원수', 'None', 'None', 'None', 'None', '평균근속년수', '연급여총액', '1인평균'],
         ['구분', '상시종업원수', 'None', 'None', 'None', 'None', '평균근속년수', '연급여총액', '1인평균'], 
         ['번호', '구분', '사업자번호', '사업장명', '주생산품', '주소', '전화번호'],
         ['설비명', 'wildcard', 'None', 'None', 'wildcard', 'None', 'None', '2015-12-31 00:00:00', 'None', 'None'],
         ['설비명', 'wildcard', 'None', 'None', 'wildcard', 'None', 'None', '2016-12-31 00:00:00', 'None', 'None'],
         ['설비명', 'wildcard', 'None', 'None', 'wildcard', 'None', 'None', '2017-12-31 00:00:00', 'None', 'None'],
         ['설비명', 'wildcard', 'None', 'None', 'wildcard', 'None', 'None', '2018-12-31 00:00:00', 'None', 'None'],
         ['설비명', 'wildcard', 'None', 'None', 'wildcard', 'None', 'None', '2019-12-31 00:00:00', 'None', 'None'],
         ['설비명', 'wildcard', 'None', 'None', 'wildcard', 'None', 'None', '2020-12-31 00:00:00', 'None', 'None'], 
         ['사업부문', '주요제품(품목)', '매출액구성비(%)'],['사업부문', '주요제품(품목)', '매출액구성비(%)'], 
         ['사업부문', '주요제품(품목)', '매출액구성비(%)'],['사업부문', '주요제품(품목)', '매출액구성비(%)'],
         ['사업부문', '주요제품(품목)', '매출액구성비(%)'],['사업부문', '주요제품(품목)', '매출액구성비(%)'],
         ['기업명', '사업자번호', '대표자명', '거래비중', '결산년도', '자본금', '자산총계', '매출액', '순이익'],
         ['기업명', '사업자번호', '대표자명', '거래비중', '결산년도', '자본금', '자산총계', '매출액', '순이익'],
         ['기업명', '사업자번호', '대표자명', '거래비중', '결산년도', '자본금', '자산총계', '매출액', '순이익'],
         ['기업명', '사업자번호', '대표자명', '거래비중', '결산년도', '자본금', '자산총계', '매출액', '순이익'],
         ['기업명', '사업자번호', '대표자명', '거래비중', '결산년도', '자본금', '자산총계', '매출액', '순이익'],
         ['기업명', '사업자번호', '대표자명', '거래비중', '결산년도', '자본금', '자산총계', '매출액', '순이익'],
         ['기업명', '사업자번호', '대표자명', '거래비중', '결산년도', '자본금', '자산총계', '매출액', '순이익'],
         ['기업명', '사업자번호', '대표자명', '거래비중', '결산년도', '자본금', '자산총계', '매출액', '순이익'],
         ['기업명', '사업자번호', '대표자명', '거래비중', '결산년도', '자본금', '자산총계', '매출액', '순이익'],
         ['기업명', '사업자번호', '대표자명', '거래비중', '결산년도', '자본금', '자산총계', '매출액', '순이익'],
         ['기업명', '사업자번호', '대표자명', '거래비중', '결산년도', '자본금', '자산총계', '매출액', '순이익'],
         ['기업명', '사업자번호', '대표자명', '거래비중', '결산년도', '자본금', '자산총계', '매출액', '순이익'],
         ['등급', '평가(산출)일자', '재무기준일자', '등급구분'],
         ['계정명', 'wildcard', 'wildcard', '2015-12-31 00:00:00'],
         ['계정명', 'wildcard', 'wildcard', '2016-12-31 00:00:00'],
         ['계정명', 'wildcard', 'wildcard', '2017-12-31 00:00:00'],
         ['계정명', 'wildcard', 'wildcard', '2018-12-31 00:00:00'],
         ['계정명', 'wildcard', 'wildcard', '2019-12-31 00:00:00'],
         ['계정명', 'wildcard', 'wildcard', '2020-12-31 00:00:00'],
         ['계정명', 'wildcard', 'wildcard', '2015-12-31 00:00:00'],
         ['계정명', 'wildcard', 'wildcard', '2016-12-31 00:00:00'],
         ['계정명', 'wildcard', 'wildcard', '2017-12-31 00:00:00'],
         ['계정명', 'wildcard', 'wildcard', '2018-12-31 00:00:00'],
         ['계정명', 'wildcard', 'wildcard', '2019-12-31 00:00:00'],
         ['계정명', 'wildcard', 'wildcard', '2020-12-31 00:00:00'],
         ['계정명', 'wildcard', 'wildcard', '2015-12-31 00:00:00'],
         ['계정명', 'wildcard', 'wildcard', '2016-12-31 00:00:00'],
         ['계정명', 'wildcard', 'wildcard', '2017-12-31 00:00:00'],
         ['계정명', 'wildcard', 'wildcard', '2018-12-31 00:00:00'],
         ['계정명', 'wildcard', 'wildcard', '2019-12-31 00:00:00'],
         ['계정명', 'wildcard', 'wildcard', '2020-12-31 00:00:00']
        ]
